from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook
from zenrows import ZenRowsClient
from shutil import copyfile
import openpyxl
import math
import requests
import datetime
import pandas as pd
import json
import os
from datetime import date
import time
import traceback
page_url = 'https://caseparts.com/'
product_json_final = []
extract_data = ""
matched_cate = ""
man = ""

def data_to_database(product_json_final,matched_cate):
    list_length = len(product_json_final)
    if list_length>0:
        print(len(product_json_final))
        input("NNNNNNNNn")
        if len(product_json_final)<51:
            dataj={
                "portal": "Caseparts",
                "category": matched_cate,
                "products":product_json_final
                }
            json_data = dataj
            print(json_data)
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Bearer RAP@BOT',
            }
            # response = requests.post('http://206.189.137.81:5052/api/v1/raw-data/', headers=headers, json=json_data)
            response = requests.post('https://hussmannxref.rapidautomation.ai/api/v1/raw-data', headers=headers, json=json_data)
            print("--",response)
            # print(response.raise_for_status())
        else:
            divisor = math.ceil(list_length)/50
            divisor = int(divisor)
            start = 0
            end=50
            for i in range(0,divisor):
                new_list = product_json_final[start:end]
                start +=50
                end+=50
                dataj={
                "portal": "Caseparts",
                "category": matched_cate,
                "products":new_list
                }
            json_data = dataj
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Bearer RAP@BOT',
            }
            # response = requests.post('http://206.189.137.81:5052/api/v1/raw-data/', headers=headers, json=json_data)
            response = requests.post('https://hussmannxref.rapidautomation.ai/api/v1/raw-data', headers=headers, json=json_data)
            print("---",response)
            # print(response.raise_for_status())
        print(len(dataj["products"]), ' data inserted to database')


def write_into_excel(folder_name, json_data):
    try:
        os.makedirs(folder_name, exist_ok=True)
        file_name = f'{folder_name}_data.xlsx'
        try:
            wb = load_workbook(file_name)
        except:
            wb = Workbook()
            sheet = wb.active
            sheet.cell(1, 1).value = 'portal'
            sheet.cell(1, 2).value = 'category'
            sheet.cell(1, 3).value = 'title'
            sheet.cell(1, 4).value = 'product-id'
            sheet.cell(1, 5).value = 'url'
            sheet.cell(1, 6).value = 'manufacturer'
            sheet.cell(1, 7).value = 'main-description'
            sheet.cell(1, 8).value = 'content'
            wb.save(file_name)
            wb.close()
            wb = load_workbook(file_name)
        sheet = wb.active
        max_row = sheet.max_row + 1
      
        wb.save(file_name)
        wb.close()
        copyfile(file_name, os.path.join(folder_name, file_name))
        
    except Exception as e:
        traceback.print_exc()
        

service = Service('path/to/chromedriver.exe')  # Replace with the actual path to chromedriver.exe
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# Navigate to the page
driver.get(page_url)
driver.maximize_window()

# Wait for the page to load
driver.implicitly_wait(12)

result = driver.page_source
excel = r"D:\\New folder\\Hussmann\\satya\\caseparts_master.xlsx"
elements = driver.find_elements(By.XPATH, "//div[@id='root']/div/div[2]/div[3]/div")
click_inside_the_matched_categories = driver.find_elements(By.XPATH, "//div[@class='sc-eHgmQL gmKbQu']/div/a")

if __name__ == "__main__":

    Manufacturers =  ["Master-Bilt"]
    # portal_Manufacturers =  ["Master-Bilt"]
    # # Manufacturers =  ["Master – Bilt","Arneg USA","Hill Phoenix","Tyler","Heatcraft","TrueMfg","Hoshizaki","Barker","Structural Concepts","Kysor Warren","Zero-Zone","Federal Industries"]
    # # portal_Manufacturers =  ["Zero Zone","Master-Bilt","Tyler","Federal","True"]
    # parts  =["curtain","cutter","fence","holder","mirror","damper","grill","honeycomb","louver","sweep","ballast","battery","block","board","breaker","cable","capacitor","connector","contactor","controller",'cord',"diffuser"]
    # parts  =["diode","drive","driver","enclosure","fixture","fuse","gauge","harness","heater","indicator","jumper","lamp","led","leveler","monitor","power supply","receptacle","relay","sensor","shield","socket","starter"]
    # parts  = ["switch","thermometer","thermostat","time clock","transducer","transformer","closer","door","frame","gasket","glide","grip",'handle',"harness-door","lift","pin","strut","track","wiper","clamp","cylinder","glass"]
    # parts  =["hinge","multiglass","plexiglass","adapter","bolt","collar","coupling","drain","fitting","hose","key","magnet","plug","post","pull","rod","spacer","strainer","stud","tool","trap","elbow"]
    # parts  = ["insert","insulation","module","tape","bracket-motor","fan","fan blade","guard","motor","motor assy","ring-motor","cover","divider","end","pan","plate","splashguard","tray","rack","cap tube","coil","compressor"]
    parts  = ["condenser","condensing unit","cooler","core","distributor","drier","evaporator","filter","orifice","pump","receiver","regulator","sight glass",'solenoid',"tank","tube",'valve',"valve-ball","valve-check","valve-solenoid","harness-shelf","product stop"]
    # parts  =["ptm","shelf","bracket","bumper",'cap',"channel",'clip',"corner","display","extrusion","joint","leg","moulding","mount","rail","retainer","support","trim"]
   
    # df = pd.read_excel("D:\\New folder\\Hussmann\\satya\\caseparts_master.xlsx",sheet_name='Sheet1')
    # column_name = 'Manufacturer'
    # column_name_part = 'Parts'
    
    # Manufacturers = df[column_name].tolist()
    # Manufacturers = Manufacturers[:13]
    # print(Manufacturers,"MANUFACTURES")
    
    # # #Sub-Zero
    
    # parts = df[column_name_part].tolist()
    
    excel_categories = parts
    for j in parts:
        if "nan" == str(j):
           
            parts.remove(j)
    excel_categories = parts
    print(excel_categories,"EXCEL CATEGORIES")
    input("****************************")
  
    today_date = date.today()
    # import datetime
    source_folder = "D:\\New folder\\Hussmann\\satya"

   
    current_date = datetime.date.today()
    
    day_name = current_date.strftime("%A")
    print("Today is", day_name)

    workbook = Workbook()
    sheet_botrun = workbook.active
    sheet_botrun['A1'].value = 'Manufacturer'
    sheet_botrun['B1'].value = 'Part Name'
    sheet_botrun['C1'].value = 'Day'
    sheet_botrun['D1'].value = 'Day Reference'
    sheet_botrun['E1'].value = 'Manufacturer Reference'
    
    bot_run_file_name = source_folder+'\\bot run report '+str(today_date)+'.xlsx'
    yesterday = today_date-datetime.timedelta(days = 1)
   
    yesterday_bot_run_file_name = source_folder+'\\bot run report '+str(yesterday)+'.xlsx'
    
   
    # input()
    last_day = 0
    last_manufacturer = 0
    if os.path.exists(yesterday_bot_run_file_name):
        print(yesterday_bot_run_file_name)
        
        wb = openpyxl.load_workbook(yesterday_bot_run_file_name)
        wb_sheet = wb.active
        last_day =  wb_sheet['D2'].value
        last_manufacturer = wb_sheet['E2'].value
        print(last_manufacturer)
        input("WWWWWWWWWWWWWWWWWWWWWWWWWWW")
        print("last day:",last_day)
    for k in Manufacturers:
        print(k)
        if "nan" == str(k):

            print("Incorrect")
            Manufacturers.remove(k)
   
    if last_day is 7:
        last_day=0
       

    if last_manufacturer is 13:
        last_manufacturer=1
    print(last_manufacturer,")))))))))))")
    # Manufacturer_to_run = Manufacturers[last_manufacturer-1]
    Manufacturer_to_run = Manufacturers[0]
    
    print(Manufacturer_to_run,"----------------")
  
    sheet_botrun['A2'].value = Manufacturer_to_run
    sheet_botrun['B2'].value = today_date
    sheet_botrun['C2'].value = today_date
    sheet_botrun['D2'].value = last_day+1
    sheet_botrun['E2'].value = last_manufacturer+1

    print('done')
        

    row_botrun = sheet_botrun.max_row
    
    workbook.save(os.path.join(bot_run_file_name))
    # day_name = 1
    print(row_botrun)
    new_date = {}
    

   
    each_manufacturer = str(Manufacturer_to_run)

    if True:        
        workbook.save(os.path.join(bot_run_file_name))
        slice_list_flag = last_day*22
        print(slice_list_flag)
        parts = parts[slice_list_flag-22:slice_list_flag]
        print(parts)
for element in range(len(elements)):
    print("eeee",element)
    try:
        time.sleep(2)

        driver.refresh
        competitor_column = driver.find_elements(By.XPATH, f"//div[@id='root']/div/div[2]/div[3]/div[{element + 1}]")
        print("compp",competitor_column)
        for i in competitor_column:
            print("iii",i)
            print(i.text)
          
            comp_inside_list = i
            comp_text = i.text
            print(comp_text,"COMPETITOR VALUE")
            
            print(Manufacturer_to_run,comp_text)
            
            # if Manufacturer_to_run == "Master – Bilt" :
            #     comp_text = Manufacturer_to_run
            if comp_text == "Federal" :
             
                man=Manufacturer_to_run.replace("Federal Industries","Federal")
                print(man)
                # comp_text = man
                
             
            elif comp_text == "Zero Zone" :
                man=Manufacturer_to_run.replace("Zero-Zone", "Zero Zone")
                # comp_text = man
            elif comp_text == "True" :
                man=Manufacturer_to_run.replace("TrueMfg", "True")
                # comp_text = man
           
            elif comp_text == "Master-Bilt":
                man=Manufacturer_to_run.replace("Master – Bilt", "Master-Bilt")
                # comp_text = man
                print("HELLO",comp_text,Manufacturer_to_run)
            elif comp_text == "Tyler" :
                man=Manufacturer_to_run.replace("Tyler", "Tyler")
                # comp_text = man
            
            print("Match",comp_text == man,"::",comp_text,"::",man)   
            
            if comp_text == man :
                print(comp_text,"CHECK COMPETITOR")
                i.click()
                print(comp_text,man,".................")
                time.sleep(3)
                click_inside_the_matched_categories = driver.find_elements(By.XPATH, "//div[@class='sc-eHgmQL gmKbQu']/div/a")
                print(click_inside_the_matched_categories,"CLICKED MATCHED CATEGORY")
                try:
                    for click_inside_the_matched_categories in range(len(click_inside_the_matched_categories)):
                        row_botrun = row_botrun+1
                        sheet_botrun.cell(row = row_botrun, column = 1 ).value =   each_manufacturer     

                        sheet_botrun.cell(row = row_botrun, column = 2 ).value =   click_inside_the_matched_categories
                        day = 1
                        sheet_botrun.cell(row = row_botrun, column = 3 ).value =   day
                       
                        print(click_inside_the_matched_categories)
                        b = driver.find_elements(By.XPATH, f"//div[@class='sc-eHgmQL gmKbQu']/div/a[{click_inside_the_matched_categories+1}]")
                        
                        try:
                            for j in b:
                                j.click()
                                time.sleep(3)
                                desc_len = len(driver.find_elements(By.XPATH, f"//tr[@class='partTableRow']//td[2]"))
                                
                                for desc_line in range(desc_len):
                                    extract_data = ""
                                    matched_cate = ""
                                    try:
                                        print("ITERATION:",desc_line)
                                        if desc_line != 0:
                                            desc_iter = desc_line -1
                                        else:
                                            desc_iter = desc_line
                                        try:
                                            part_number = driver.find_elements(By.XPATH,f"//tr[@class='partTableRow']//td[1]//a")[desc_iter]
                                            part_number.click()
                                            print("PART NUMBER CLICKED")
                                        except:
                                            part_number = driver.find_elements(By.XPATH,f"//tr[@class='partTableRow']//td[2]//a")[desc_iter]
                                            part_number.click()
                                            time.sleep(5)
                                            print("EXCEPTIPON CLICKED")
                                        
                                        # desc = driver.find_elements(By.XPATH,f"//*[@class='sc-tilXH dPblDn']")[0].text---ORIGINAL
                                        # descc = driver.find_elements(By.XPATH,f"//*[@class='sc-ktHwxA euZMGQ']")[0].text----29/07/2024
                                        descc = driver.find_elements(By.XPATH,f"//*[@class='sc-ktHwxA hsIPbh']")[0].text
                                        desc = descc.replace("'","").replace('"','')
                                        print(desc,"EXTRACTED DESCRIPTION")
                                        
                                        # content_path =driver.find_elements(By.XPATH,f"//*[@class='sc-gisBJw flQCao']")[0].text------29/07/2024
                                        content_path =driver.find_elements(By.XPATH,f"//*[@class='sc-fAjcbJ eGiVfx']")[0].text
                                        print(content_path,"CONTENT PATH")
                                        content_path_list = [content_path]
                                       
                                        content_path_list[0] = content_path_list[0].replace('\nDisplayMfg:Federal', '')
                                        print(content_path_list[0],"content path list ")
                                       
                                      
                                        
                                        split_desc = desc.split(" ")

                                        split_desc.reverse()
                                        print(split_desc,"SPLIT DESC")
                                        for desc_list in split_desc:
                                            
                                            if desc_list.lower().strip() in [x.lower() for x in excel_categories]:

                                        

                                                current_url = driver.current_url
                                                print(current_url)
                                                # extract_data = driver.find_elements(By.XPATH,f"//*[@class='sc-gGBfsJ bCeYkR']")[0].text-----ORIGINAL
                                                #extract_data = driver.find_elements(By.XPATH,f"//*[@class='sc-fYxtnH iakXjq']")[0].text------29/07/2024
                                                extract_data = driver.find_elements(By.XPATH,f"//*[@class='sc-jtRfpW DUheu']")[0].text
                                                
                                                print(extract_data,"EXTRACT DATA")
                                                matched_cate = desc_list.lower()
                                                print(matched_cate, desc,'---------------------MATCHING')
                                              
                                                break


                                            
                                        
                                        if extract_data != "" and matched_cate != "":
                                            data = extract_data.split('\n')
                                            result = {}
                                    
                                            folder_path = rf'D:\New folder\Hussmann\json\{comp_text}' + "\\"
                                            
                                            os.makedirs(folder_path, exist_ok= True)
                                    
                                            for item in data:
                                                if item.startswith('Part #:'):
                                                    part_num = item.split(': ')[1]
                                                if item.startswith('DisplayMfg:'):
                                                    print(item)
                                               
                                                    if ':' in item:
                                                        manufacture = item.split(':')[1]
                                                        print(manufacture,"......................................")

                                                       
                                                       
                                                        current_date = datetime.date.today()

                                                        formatted_date = current_date.strftime("%d_%b_%Y")
                                                        
                                                        portal_name="caseparts_"
                                                        #creating json name 
                                                        Json_name = portal_name+matched_cate+'_'+formatted_date+'_set1'
                                                        
                                                        file_path = os.path.join(folder_path,f"{Json_name}.json")
                                                      
                                                        
                                                        
                                            
                                            con_path = content_path_list[0].replace("\n",",")
                                            print(con_path,"CON PATH")
                                            data_dict = dict(item.replace('//\//', '').replace('"',"").split(":") for item in con_path.split(",") if ":" in item)
                                            print(data_dict,"DATA DICT")
                                            
                                           
                                            
                                            keys_to_remove = ["Manufacturer", "DisplayMfg"]
                                            print(keys_to_remove,"KEYS TO REMOVE")
                                            output_content = "NA"
                                            for key in keys_to_remove:
                                                if key in data_dict:
                                                    del data_dict[key]
                                            
                                            if type(data_dict) != dict:
                                                DDDDD = json.loads(data_dict)
                                                print(DDDDD,"aaaaaaa")
                                            
                                            if data_dict == {}:
                                                output_content = 'N/A'
                                                print(output_content,"Bbbbbb")
                                                
                                            else:
                                                output_content = data_dict
                                                print(output_content,"CCCCCCCC")
                                           
                                            keys_to_remove = ["Manufacturer", "DisplayMfg"]
                                            
                                            try:
                                                # print(part_num,current_url,matched_cate,desc,part_num,manufacture,output_content)
                                                # input()
                                                if part_num!='N/A':
                                                    print(part_num,"PART NUMBER")
                                                    product_json = {}
                                                    
                                                    product_json={

                                    
                                                                    "recognized_attributes": {
                                                                        "sku": part_num,
                                                                        "url": current_url,
                                                                        "title": matched_cate,
                                                                        "main-description": desc,
                                                                        "product-id": part_num,
                                                                        "oem-product-id": part_num,
                                                                        "compliant-with": "N/A",
                                                                        # "manufacturer": manufacture.lower(),
                                                                        "manufacturer": "master-bilt",
                                                                    },
                                                                    "all_extracted_data": [
                                                                        {
                                                                            "extraction_priority": 1,
                                                                            "section_name": "details",
                                                                            "media_type": "text",
                                                                            "format_type": "json",
                                                                            "content": output_content,
                                                                        }
                                                                    ]
                                                                }
                                                    print(product_json)
                                                    product_json_final.append(product_json)
                                                
                                            except Exception as e:
                                                # print(e)
                                                traceback.print_exc()
                                              
                                                # input("---------")
                                            if os.path.isfile(file_path):
                                                with open(file_path,"r") as filedata:
                                                    old_json  = json.load(filedata)
                                                
                                                old_json.append(product_json_final)
                                                save_file = open(file_path, "w")  
                                                json.dump(old_json, save_file)  
                                                save_file.close()
                                            # else:
 
                                            with open(file_path, 'w') as json_file:
                                                
                                                json.dump(product_json_final, json_file)
                                             
                                            write_into_excel(comp_text, product_json_final)
                                           
                                        
                                        else:
                                            print("---No categaory matching-------")
                                    
                                        
                                    except Exception as e:
                                        print(e)
                                        # input()
                            
                                        
                            show_all_competitors = driver.find_elements(By.XPATH,f"//*[contains(text(),'Show All Pages')]")[0]
                            show_all_competitors.click() 
                            time.sleep(5)
                            
                            try :
                                manufacture_headers_from_website = driver.find_elements(By.XPATH, f"//div[@id='root']/div/div[2]/div[3]/div[{element+1}]")

                                for comp_inside_list_t in manufacture_headers_from_website:
                                   
                                        comp_text_t = comp_inside_list_t.text
                                        print(comp_text_t)
                                        
                                        if comp_text_t in Manufacturer_to_run:

                                                    
                                            comp_inside_list_t.click()
                            except Exception as e:
                                print(e)
                                print("compppp")
                            
                        except Exception as e:
                            print(e)
                            print("loop")
                    
                except Exception as e:
                    print(e)
                    print("loop out")
                    # input()
    except Exception as e:
        print(e) 
        traceback.print_exc()
        # 
        # ("---failed---")
print("----------------------------------------------------------------------")
print(product_json_final)
print(type(product_json_final),">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
print(matched_cate)
print(parts)

for each_part in parts:
    print(each_part, "MMMM<<<<<<<<<<<<<<<<<<<<<<<<")
    list_parts_json = []
    for dict_json in product_json_final:
        title_check = dict_json["recognized_attributes"]["title"]
        print(title_check)
        if title_check.lower() == each_part.lower():
            list_parts_json.append(dict_json)
    print(len(list_parts_json))        
    if len(list_parts_json) != 0:
        print("pushed data"," &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
        print(list_parts_json)
        data_to_database(list_parts_json,each_part)
 



    

    
















